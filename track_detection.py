import cv2
import numpy as np
import os
import time
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox

try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except Exception:
    EXCEL_AVAILABLE = False

EXCEL_FILE = "track_results.xlsx"
IMAGE_FOLDER = "images"
EDGE_THRESHOLD = 3000

COLOR_BG = "#0c101e"
COLOR_PANEL = "#151b2d"
COLOR_PANEL_ALT = "#1b2236"
COLOR_BORDER = "#2a3552"
COLOR_TEXT = "#e6eaf2"
COLOR_MUTED = "#9aa6bf"
COLOR_OK = "#33c26b"
COLOR_WARN = "#e04f5f"
COLOR_ACCENT = "#4aa3ff"
COLOR_BUTTON = "#2b3552"
COLOR_BUTTON_HOVER = "#39466a"
COLOR_BUTTON_STOP = "#b73846"
COLOR_BUTTON_STOP_HOVER = "#c84b59"
COLOR_BUTTON_START = "#2fae5d"
COLOR_BUTTON_START_HOVER = "#3ac26d"
COLOR_LIVE_BG = "#2fae5d"

FONT_BASE = ("Segoe UI", 11)
FONT_HEADER = ("Segoe UI", 13, "bold")
FONT_TITLE = ("Segoe UI", 15, "bold")
FONT_BUTTON = ("Segoe UI", 12, "bold")
FONT_MONO = ("Consolas", 10)


def _init_excel():
    if not EXCEL_AVAILABLE:
        return None, None

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Image Name", "Detection Result"])
        wb.save(EXCEL_FILE)
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    return wb, ws


def detect_crack(image_path):
    """Detect if the railway track has a crack using edge analysis from image file."""
    image = cv2.imread(image_path)
    if image is None:
        return "Invalid Image"

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blurred, 50, 150)

    edge_count = cv2.countNonZero(edged)

    if edge_count > EDGE_THRESHOLD:
        return "crack"
    return "no crack"


def detect_crack_from_frame(frame):
    """Detect if the live camera frame has a crack using edge analysis."""
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blurred, 50, 150)

    edge_count = cv2.countNonZero(edged)

    if edge_count > EDGE_THRESHOLD:
        return "crack", edged
    return "no crack", edged


def process_all_images():
    wb, ws = _init_excel()
    if wb is None or ws is None:
        print("openpyxl not installed. Skipping Excel logging.")
        return

    if not os.path.isdir(IMAGE_FOLDER):
        print("Image folder not found.")
        return

    for file_name in os.listdir(IMAGE_FOLDER):
        if file_name.lower().endswith((".jpg", ".jpeg", ".png")):
            full_path = os.path.join(IMAGE_FOLDER, file_name)
            result = detect_crack(full_path)
            print(f"{file_name} -> {result}")
            ws.append([file_name, result])

    wb.save(EXCEL_FILE)
    print("Results have been recorded in 'track_results.xlsx'")


class TrackDetectionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Track Crack Detection")
        self.root.minsize(1100, 650)
        self.root.geometry("1280x740")
        self.root.configure(bg=COLOR_BG)

        self.cap = None
        self.running = False
        self.last_log_time = 0.0
        self.last_console_time = 0.0
        self.last_result = None
        self.video_size = (720, 405)

        self._build_ui()
        self._set_idle_state()

        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_ui(self):
        self.style = ttk.Style(self.root)
        self.style.theme_use("clam")

        self.bg_canvas = tk.Canvas(self.root, bg=COLOR_BG, highlightthickness=0)
        self.bg_canvas.pack(fill="both", expand=True)

        self.content_frame = tk.Frame(self.bg_canvas, bg=COLOR_BG)
        self.content_window = self.bg_canvas.create_window(0, 0, anchor="nw", window=self.content_frame)

        self.root.bind("<Configure>", self._on_resize)

        self.content_frame.grid_rowconfigure(0, weight=4)
        self.content_frame.grid_rowconfigure(1, weight=1)
        self.content_frame.grid_columnconfigure(0, weight=1)

        top_frame = tk.Frame(self.content_frame, bg=COLOR_BG)
        top_frame.grid(row=0, column=0, sticky="nsew")
        top_frame.grid_columnconfigure(0, weight=3)
        top_frame.grid_columnconfigure(1, weight=2)
        top_frame.grid_columnconfigure(2, weight=1)
        top_frame.grid_rowconfigure(0, weight=1)

        left_panel = tk.Frame(top_frame, bg=COLOR_PANEL, highlightthickness=1, highlightbackground=COLOR_BORDER)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(20, 10), pady=20)

        header = tk.Frame(left_panel, bg=COLOR_PANEL_ALT)
        header.pack(fill="x")
        header_label = tk.Label(header, text="Detection:", bg=COLOR_PANEL_ALT, fg=COLOR_TEXT, font=FONT_HEADER)
        header_label.pack(side="left", padx=(12, 6), pady=8)
        self.detection_value = tk.Label(header, text="idle", bg=COLOR_PANEL_ALT, fg=COLOR_MUTED, font=FONT_HEADER)
        self.detection_value.pack(side="left", pady=8)

        video_border = tk.Frame(left_panel, bg=COLOR_BORDER)
        video_border.pack(fill="both", expand=True, padx=12, pady=12)

        self.video_label = tk.Label(video_border, bg="#0b0f1f", fg=COLOR_MUTED)
        self.video_label.pack(fill="both", expand=True, padx=2, pady=2)
        if not PIL_AVAILABLE:
            self.video_label.configure(text="Install Pillow to view camera feed")

        right_panel = tk.Frame(top_frame, bg=COLOR_PANEL, highlightthickness=1, highlightbackground=COLOR_BORDER)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=10, pady=20)

        right_header = tk.Label(right_panel, text="Output & Logging", bg=COLOR_PANEL_ALT, fg=COLOR_TEXT, font=FONT_HEADER, anchor="w")
        right_header.pack(fill="x", padx=0, pady=0)
        right_header.configure(padx=12, pady=8)

        right_body = tk.Frame(right_panel, bg=COLOR_PANEL)
        right_body.pack(fill="both", expand=True, padx=14, pady=10)

        self.result_value = self._stat_row(right_body, "Result", "-", COLOR_OK)
        self.timestamp_value = self._stat_row(right_body, "Timestamp", "-", COLOR_TEXT)
        self.edge_value = self._stat_row(right_body, "Edge Avg Intensity", "-", COLOR_TEXT)

        log_label = tk.Label(right_body, text="Log:", bg=COLOR_PANEL, fg=COLOR_TEXT, font=FONT_HEADER)
        log_label.pack(anchor="w", pady=(12, 6))

        self.output_log = tk.Text(right_body, height=8, bg="#0f1426", fg=COLOR_TEXT, font=FONT_MONO, bd=0, wrap="word")
        self.output_log.pack(fill="both", expand=True)
        self.output_log.configure(state="disabled")

        control_panel = tk.Frame(top_frame, bg=COLOR_BG)
        control_panel.grid(row=0, column=2, sticky="nsew", padx=(10, 20), pady=20)
        control_panel.grid_rowconfigure(0, weight=1)

        buttons_frame = tk.Frame(control_panel, bg=COLOR_BG)
        buttons_frame.pack(fill="y", expand=True)

        self.start_button = self._make_button(buttons_frame, "Start", COLOR_BUTTON_START, self.start_capture, COLOR_BUTTON_START_HOVER)
        self.stop_button = self._make_button(buttons_frame, "Stop", COLOR_BUTTON_STOP, self.stop_capture, COLOR_BUTTON_STOP_HOVER)
        self.settings_button = self._make_button(buttons_frame, "Settings", COLOR_BUTTON, self.open_settings, COLOR_BUTTON_HOVER)
        self.exit_button = self._make_button(buttons_frame, "Exit", COLOR_BUTTON, self.on_close, COLOR_BUTTON_HOVER)

        bottom_panel = tk.Frame(self.content_frame, bg=COLOR_PANEL, highlightthickness=1, highlightbackground=COLOR_BORDER)
        bottom_panel.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
        bottom_panel.grid_columnconfigure(0, weight=1)

        bottom_header = tk.Label(bottom_panel, text="System Log", bg=COLOR_PANEL_ALT, fg=COLOR_TEXT, font=FONT_HEADER, anchor="w")
        bottom_header.pack(fill="x", padx=0, pady=0)
        bottom_header.configure(padx=12, pady=6)

        bottom_body = tk.Frame(bottom_panel, bg=COLOR_PANEL)
        bottom_body.pack(fill="both", expand=True, padx=12, pady=10)
        bottom_body.grid_columnconfigure(0, weight=1)
        bottom_body.grid_columnconfigure(1, weight=0)

        log_frame = tk.Frame(bottom_body, bg=COLOR_PANEL)
        log_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        scrollbar = ttk.Scrollbar(log_frame, orient="vertical")
        self.console_log = tk.Text(log_frame, height=6, bg="#0f1426", fg=COLOR_OK, font=FONT_MONO, bd=0, wrap="word", yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.console_log.yview)
        scrollbar.pack(side="right", fill="y")
        self.console_log.pack(fill="both", expand=True)
        self.console_log.configure(state="disabled")

        self.live_indicator = tk.Label(bottom_body, text="LIVE", bg=COLOR_PANEL_ALT, fg=COLOR_MUTED, font=FONT_TITLE, width=8, height=2)
        self.live_indicator.grid(row=0, column=1, sticky="e")

    def _stat_row(self, parent, label, value, value_color):
        row = tk.Frame(parent, bg=COLOR_PANEL)
        row.pack(fill="x", pady=4)
        label_widget = tk.Label(row, text=f"{label}:", bg=COLOR_PANEL, fg=COLOR_MUTED, font=FONT_BASE)
        label_widget.pack(side="left")
        value_widget = tk.Label(row, text=value, bg=COLOR_PANEL, fg=value_color, font=FONT_BASE)
        value_widget.pack(side="left", padx=(6, 0))
        return value_widget

    def _make_button(self, parent, text, bg, command, hover_bg):
        button = tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg,
            fg=COLOR_TEXT,
            activebackground=hover_bg,
            activeforeground=COLOR_TEXT,
            font=FONT_BUTTON,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=COLOR_BORDER,
            padx=12,
            pady=10,
        )
        button.pack(fill="x", pady=8)
        return button

    def _set_idle_state(self):
        self.detection_value.configure(text="idle", fg=COLOR_MUTED)
        self.result_value.configure(text="-", fg=COLOR_MUTED)
        self.timestamp_value.configure(text="-", fg=COLOR_TEXT)
        self.edge_value.configure(text="-", fg=COLOR_TEXT)
        self.live_indicator.configure(bg=COLOR_PANEL_ALT, fg=COLOR_MUTED)
        self.start_button.configure(state="normal")
        self.stop_button.configure(state="disabled")

    def _append_output_log(self, message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.output_log.configure(state="normal")
        self.output_log.insert("end", f"[{timestamp}] {message}\n")
        self.output_log.see("end")
        self.output_log.configure(state="disabled")

    def _append_console_log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.console_log.configure(state="normal")
        self.console_log.insert("end", f"[{timestamp}] {message}\n")
        self.console_log.see("end")
        self.console_log.configure(state="disabled")

    def _on_resize(self, event):
        if event.widget is not self.root:
            return
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        self.bg_canvas.config(width=width, height=height)
        self.bg_canvas.itemconfig(self.content_window, width=width, height=height)
        self._draw_gradient(width, height)

    def _draw_gradient(self, width, height):
        self.bg_canvas.delete("gradient")
        if width <= 1 or height <= 1:
            return
        r1, g1, b1 = (12, 16, 30)
        r2, g2, b2 = (20, 27, 46)
        for i in range(height):
            ratio = i / (height - 1)
            r = int(r1 + (r2 - r1) * ratio)
            g = int(g1 + (g2 - g1) * ratio)
            b = int(b1 + (b2 - b1) * ratio)
            color = f"#{r:02x}{g:02x}{b:02x}"
            self.bg_canvas.create_line(0, i, width, i, fill=color, tags="gradient")
        self.bg_canvas.lower("gradient")

    def start_capture(self):
        if self.running:
            return

        if self.cap is None:
            self.cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
            if not self.cap.isOpened():
                self.cap.release()
                self.cap = None
                messagebox.showerror("Camera Error", "Cannot open the webcam.")
                return

        self.running = True
        self._append_console_log("Live capture started")
        self.live_indicator.configure(bg=COLOR_LIVE_BG, fg=COLOR_TEXT)
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        self._update_frame()

    def stop_capture(self):
        if not self.running:
            return
        self.running = False
        self._append_console_log("Live capture stopped")
        self.live_indicator.configure(bg=COLOR_PANEL_ALT, fg=COLOR_MUTED)
        self.start_button.configure(state="normal")
        self.stop_button.configure(state="disabled")
        if self.cap is not None:
            self.cap.release()
            self.cap = None

    def open_settings(self):
        messagebox.showinfo("Settings", "Settings panel is not wired yet. Add thresholds or camera selection here.")

    def _update_frame(self):
        if not self.running or self.cap is None:
            return

        ret, frame = self.cap.read()
        if not ret:
            self._append_console_log("Failed to grab frame")
            self.stop_capture()
            return

        result, edged = detect_crack_from_frame(frame)
        edge_avg = float(np.mean(edged)) if edged is not None else 0.0

        result_text = "Crack Detected" if result == "crack" else "No Crack"
        result_color = COLOR_WARN if result == "crack" else COLOR_OK

        self.detection_value.configure(text=result, fg=result_color)
        self.result_value.configure(text=result_text, fg=result_color)
        self.timestamp_value.configure(text=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.edge_value.configure(text=f"{edge_avg:.1f}")

        now = time.time()
        if self.last_result != result or (now - self.last_log_time) > 2.5:
            self._append_output_log(result_text)
            self.last_result = result
            self.last_log_time = now
        if now - self.last_console_time > 1.5:
            self._append_console_log("Frame successfully captured")
            self.last_console_time = now

        if PIL_AVAILABLE:
            display_frame = cv2.resize(frame, self.video_size)
            overlay = display_frame.copy()
            cv2.putText(
                overlay,
                f"Detection: {result}",
                (12, 30),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.9,
                (0, 0, 255) if result == "crack" else (0, 255, 0),
                2,
            )
            rgb = cv2.cvtColor(overlay, cv2.COLOR_BGR2RGB)
            image = Image.fromarray(rgb)
            photo = ImageTk.PhotoImage(image=image)
            self.video_label.configure(image=photo, text="")
            self.video_label.image = photo

        self.root.after(30, self._update_frame)

    def on_close(self):
        if self.running:
            self.stop_capture()
        if self.cap is not None:
            self.cap.release()
        self.root.destroy()


def main():
    root = tk.Tk()
    app = TrackDetectionApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
